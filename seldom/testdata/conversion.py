"""
Data type conversion of different files
"""
import csv
import codecs
from itertools import islice
try:
    from openpyxl import load_workbook
except ModuleNotFoundError:
    raise ModuleNotFoundError("Please install the library. https://pypi.org/project/openpyxl/")


def csv_to_list(file=None, line=1):
    """
    Convert CSV file data to list
    :param file: Path to file
    :param line: Start line of read data
    :return:  list data

    @data(csv_to_list(/home/user/seldom/data.csv, line=1))
    def test_login(self, username, password):
        print(username)
        print(password)
    """
    if file is None:
        raise FileExistsError("Please specify the CSV file to convert.")

    table_data = []
    csv_data = csv.reader(codecs.open(file, 'r', 'utf_8_sig'))
    for line in islice(csv_data, line - 1, None):
        table_data.append(line)

    return table_data


def excel_to_list(file=None, sheet="Sheet1", line=1):
    """
    Convert CSV file data to list
    :param file: Path to file
    :param sheet: excel sheet, default name is Sheet1
    :param line: Start line of read data
    :return: list data

    @data(excel_to_list(/home/user/seldom/data.xlsx, sheet="Sheet1", line=1))
    def test_login(self, username, password):
        print(username)
        print(password)
    """
    if file is None:
        raise FileExistsError("Please specify the Excel file to convert.")

    excel_table = load_workbook(file)
    sheet = excel_table[sheet]

    table_data = []
    for line in sheet.iter_rows(line, sheet.max_row):
        line_data = []
        for field in line:
            line_data.append(field.value)
        table_data.append(line_data)

    return table_data
